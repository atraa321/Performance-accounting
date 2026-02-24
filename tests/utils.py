from __future__ import annotations
from typing import Dict, List
import openpyxl

REQUIRED_FIELDS = ["name","role","direct_total","pool_nursing","pool_doctor","surplus","grand_total"]

def read_gold_standard(gold_excel_path: str) -> List[Dict]:
    wb = openpyxl.load_workbook(gold_excel_path, data_only=True)
    ws = wb["金标准结果"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        rec = dict(zip(headers, r))
        rows.append({
            "name": rec["姓名"],
            "role": rec["岗位"],
            "direct_total": float(rec["DirectPay合计"]),
            "pool_nursing": float(rec["护理池分配"]),
            "pool_doctor": float(rec["医师池分配"]),
            "surplus": float(rec["科室盈余"]),
            "grand_total": float(rec["最终应发合计"]),
        })
    return rows

def normalize_rows(rows: List[Dict]) -> List[Dict]:
    out = []
    for row in rows:
        missing = [k for k in REQUIRED_FIELDS if k not in row]
        if missing:
            raise AssertionError(f"Summary row missing fields: {missing}. Got keys={list(row.keys())}")
        out.append({
            "name": str(row["name"]),
            "role": str(row["role"]),
            "direct_total": float(row["direct_total"]),
            "pool_nursing": float(row["pool_nursing"]),
            "pool_doctor": float(row["pool_doctor"]),
            "surplus": float(row["surplus"]),
            "grand_total": float(row["grand_total"]),
        })
    return out

def index_by_name(rows: List[Dict]) -> Dict[str, Dict]:
    return {r["name"]: r for r in rows}

def assert_close(a: float, b: float, tol: float = 0.01):
    # 金额到分：默认容忍 1 分
    if abs(a - b) > tol:
        raise AssertionError(f"Amount mismatch: got {a:.2f}, expected {b:.2f}, tol={tol}")
