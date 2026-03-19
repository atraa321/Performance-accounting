from __future__ import annotations

import io
import tempfile
from typing import Optional
from decimal import Decimal
import os

from datetime import datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, select

from app.api.deps import get_db
from app.calc.engine import calculate_run
from app.calc.importer import import_excel
from app.calc.night_schedule import parse_night_shift_counts
from app.calc.validator import validate_run_data, save_validation_results_to_qc
from app.calc.utils import normalize_item_name, extract_period_tag, to_decimal, round_money
from app.core.audit import get_audit_logger
from app.core.paths import TMP_DIR
from app.models.models import (
    DictItemBehavior,
    DictItemMapping,
    DimEmployeeMonth,
    FactPayDetail,
    FactPool,
    FactPoolAlloc,
    FactPaySummary,
    QcIssue,
    RawDoctorWorkload,
    RawHospitalPerfItem,
    RawManualDoctorWorkloadPay,
    RawManualEntry,
    RawManualPoolAdjust,
    RawManualStudyLeavePay,
    RawNightShift,
    RawNurseWorkload,
    RawReadingFee,
    RawRoster,
    ReconcileItem,
    RunBatch,
)
from app.schemas.run import (
    ManualEntryPayload,
    ManualEntryPayloadV2,
    RunCreate,
    RunResponse,
    SummaryResponse,
)

router = APIRouter()

PDF_HOSPITAL_NAME = "平顶山市第五人民医院"


def _register_pdf_font() -> str:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_candidates = [
        (r"C:\Windows\Fonts\msyh.ttc", "MicrosoftYaHei", {"subfontIndex": 0}),
        (r"C:\Windows\Fonts\msyh.ttf", "MicrosoftYaHei", {}),
        (r"C:\Windows\Fonts\simsun.ttc", "SimSun", {"subfontIndex": 0}),
        (r"C:\Windows\Fonts\simsun.ttf", "SimSun", {}),
    ]
    for path, name, options in font_candidates:
        if not os.path.exists(path):
            continue
        try:
            pdfmetrics.registerFont(TTFont(name, path, **options))
            return name
        except Exception:
            continue
    return "Helvetica"

RAW_SHEET_MAP = {
    "hospital_perf_item": RawHospitalPerfItem,
    "roster": RawRoster,
    "night_shift": RawNightShift,
    "reading_fee": RawReadingFee,
    "doctor_workload": RawDoctorWorkload,
    "nurse_workload": RawNurseWorkload,
    "manual_doctor_workload_pay": RawManualDoctorWorkloadPay,
}

RAW_SHEET_LABELS = {
    "hospital_perf_item": "院发绩效表",
    "roster": "绩效发放名单",
    "night_shift": "夜班统计",
    "reading_fee": "判读费",
    "doctor_workload": "医师工作量",
    "nurse_workload": "护士工作量",
}

RUN_SCOPED_MODELS = (
    RawHospitalPerfItem,
    RawRoster,
    RawNightShift,
    RawReadingFee,
    RawDoctorWorkload,
    RawNurseWorkload,
    RawManualDoctorWorkloadPay,
    RawManualPoolAdjust,
    RawManualStudyLeavePay,
    RawManualEntry,
    DimEmployeeMonth,
    FactPool,
    FactPoolAlloc,
    FactPayDetail,
    FactPaySummary,
    ReconcileItem,
    QcIssue,
)


def _replace_manual_rows(db, run_id: int, model, rows):
    db.execute(delete(model).where(model.run_id == run_id))
    for row in rows:
        name = row.name if getattr(row, "name", None) is not None else row["name"]
        amount = row.amount if getattr(row, "amount", None) is not None else row["amount"]
        db.add(model(run_id=run_id, name=name, amount=amount))
    db.commit()


def _ensure_table(db, model):
    bind = db.get_bind()
    model.__table__.create(bind=bind, checkfirst=True)


def _delete_run_rows(db, run_id: int, models) -> None:
    for model in models:
        db.execute(delete(model).where(model.run_id == run_id))


@router.get("/runs/{run_id}/manual/workload")
def get_manual_workload(run_id: int, db=Depends(get_db)):
    rows = db.execute(
        select(RawManualDoctorWorkloadPay).where(RawManualDoctorWorkloadPay.run_id == run_id)
    ).scalars().all()
    return [{"name": r.name, "amount": float(r.amount)} for r in rows]


@router.post("/runs/{run_id}/manual/workload")
def save_manual_workload(run_id: int, payload: ManualEntryPayload, db=Depends(get_db)):
    run = db.get(RunBatch, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    _replace_manual_rows(db, run_id, RawManualDoctorWorkloadPay, payload.rows)
    return {"status": "ok", "count": len(payload.rows)}


@router.get("/runs/{run_id}/manual/study-leave")
def get_manual_study_leave(run_id: int, db=Depends(get_db)):
    _ensure_table(db, RawManualStudyLeavePay)
    rows = db.execute(
        select(RawManualStudyLeavePay).where(RawManualStudyLeavePay.run_id == run_id)
    ).scalars().all()
    return [{"name": r.name, "amount": float(r.amount)} for r in rows]


@router.post("/runs/{run_id}/manual/study-leave")
def save_manual_study_leave(run_id: int, payload: ManualEntryPayload, db=Depends(get_db)):
    run = db.get(RunBatch, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    _ensure_table(db, RawManualStudyLeavePay)
    _replace_manual_rows(db, run_id, RawManualStudyLeavePay, payload.rows)
    return {"status": "ok", "count": len(payload.rows)}


@router.get("/runs/{run_id}/manual/entries")
def get_manual_entries(run_id: int, db=Depends(get_db)):
    _ensure_table(db, RawManualEntry)
    rows = db.execute(
        select(RawManualEntry).where(RawManualEntry.run_id == run_id)
    ).scalars().all()
    return [
        {
            "target_type": r.target_type,
            "target_value": r.target_value,
            "item_type": r.item_type,
            "amount": float(r.amount),
        }
        for r in rows
    ]


@router.post("/runs/{run_id}/manual/entries")
def save_manual_entries(run_id: int, payload: ManualEntryPayloadV2, db=Depends(get_db)):
    run = db.get(RunBatch, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    _ensure_table(db, RawManualEntry)
    item_type_map = {
        "WORKLOAD": "WORKLOAD",
        "工作量": "WORKLOAD",
        "WORKLOAD_MANUAL_PAY": "WORKLOAD",
        "STUDY_LEAVE": "STUDY_LEAVE",
        "进修产假补贴": "STUDY_LEAVE",
        "STUDY_LEAVE_SUBSIDY": "STUDY_LEAVE",
        "OTHER": "OTHER",
        "其他": "OTHER",
        "MANUAL_OTHER_PAY": "OTHER",
    }
    target_type_map = {
        "PERSON": "PERSON",
        "个人": "PERSON",
        "POOL": "POOL",
        "池子": "POOL",
    }
    pool_value_map = {
        "NURSING_POOL": "NURSING_POOL",
        "护士池": "NURSING_POOL",
        "护理池": "NURSING_POOL",
        "DOCTOR_POOL": "DOCTOR_POOL",
        "医生池": "DOCTOR_POOL",
        "医师池": "DOCTOR_POOL",
    }
    db.execute(delete(RawManualEntry).where(RawManualEntry.run_id == run_id))
    for row in payload.rows:
        item_type_raw = str(row.item_type or "").strip()
        item_type = item_type_map.get(item_type_raw, item_type_raw)
        target_type_raw = str(row.target_type or "").strip()
        target_type = target_type_map.get(target_type_raw, target_type_raw)
        target_value_raw = str(row.target_value or "").strip()
        target_value = pool_value_map.get(target_value_raw, target_value_raw)
        db.add(
            RawManualEntry(
                run_id=run_id,
                target_type=target_type,
                target_value=target_value,
                item_type=item_type,
                amount=row.amount,
            )
        )
    db.commit()
    return {"status": "ok", "count": len(payload.rows)}


@router.get("/runs/{run_id}/manual/allocatable")
def get_manual_allocatable(run_id: int, db=Depends(get_db)):
    mapping_rows = (
        db.execute(
            select(DictItemMapping).where(DictItemMapping.is_active == True)
        )
        .scalars()
        .all()
    )
    mapping = [(r.raw_item_name, r.item_code, r.priority) for r in mapping_rows]
    mapping.sort(key=lambda r: r[2])

    def _map_item_code(name_norm: str) -> Optional[str]:
        for raw_name, item_code, _ in mapping:
            if raw_name in name_norm:
                return item_code
        return None

    behavior_rows = db.execute(select(DictItemBehavior)).scalars().all()
    behavior_map = {b.item_code: b.behavior_type for b in behavior_rows}

    totals = {}
    display_name_by_norm = {}
    rows = db.execute(
        select(RawHospitalPerfItem).where(RawHospitalPerfItem.run_id == run_id)
    ).scalars().all()
    for item in rows:
        item_code = _map_item_code(item.item_name_norm) or "UNCLASSIFIED_ITEM"
        name_norm = normalize_item_name(item.item_name)
        if name_norm not in display_name_by_norm or len(item.item_name) > len(display_name_by_norm[name_norm]):
            display_name_by_norm[name_norm] = item.item_name
        if item_code in ("WORKLOAD_MANUAL_PAY", "STUDY_LEAVE_SUBSIDY"):
            totals[name_norm] = totals.get(name_norm, Decimal("0")) + Decimal(str(item.amount))
            continue
        if item_code == "UNCLASSIFIED_ITEM":
            totals[name_norm] = totals.get(name_norm, Decimal("0")) + Decimal(str(item.amount))
            continue
        behavior = behavior_map.get(item_code)
        if behavior in (None, "UNCLASSIFIED", "RECON_ONLY"):
            totals[name_norm] = totals.get(name_norm, Decimal("0")) + Decimal(str(item.amount))

    items = [
        {"item_name": display_name_by_norm.get(name, name), "amount": float(amount)}
        for name, amount in totals.items()
    ]
    items.sort(key=lambda x: x["item_name"])
    return {"items": items}


@router.post("/runs", response_model=RunResponse)
def create_run(payload: RunCreate, db=Depends(get_db)):
    run = RunBatch(
        month=payload.month,
        dept_name=payload.dept_name,
        rule_version=payload.rule_version,
        status="DRAFT",
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    return RunResponse(
        run_id=run.id,
        month=run.month,
        dept_name=run.dept_name,
        rule_version=run.rule_version,
    )


@router.get("/runs")
def list_runs(db=Depends(get_db)):
    rows = db.execute(select(RunBatch)).scalars().all()
    return [
        {
            "id": r.id,
            "month": r.month,
            "dept_name": r.dept_name,
            "rule_version": r.rule_version,
            "status": r.status,
            "created_at": r.created_at,
            "locked_at": r.locked_at,
        }
        for r in rows
    ]


@router.post("/runs/{run_id}/lock")
def lock_run(run_id: int, db=Depends(get_db)):
    run = db.get(RunBatch, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    run.status = "LOCKED"
    run.locked_at = datetime.utcnow()
    db.commit()
    get_audit_logger(db).log_run_locked(run_id)
    return {"status": "locked"}


@router.post("/runs/{run_id}/import/excel")
def import_excel_api(run_id: int, clear_existing: bool = True, file: UploadFile = File(...), db=Depends(get_db)):
    run = db.get(RunBatch, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    audit_logger = get_audit_logger(db)
    if clear_existing:
        _delete_run_rows(db, run_id, RUN_SCOPED_MODELS)
        db.commit()
    tmp_path = None
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=TMP_DIR) as tmp:
        content = file.file.read()
        tmp.write(content)
        tmp_path = tmp.name
    try:
        result = import_excel(db, run_id, tmp_path)
    except Exception as exc:
        audit_logger.log_error("DATA_IMPORT", "导入Excel数据", str(exc), run_id=run_id)
        raise
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    result["cleared_existing"] = clear_existing
    audit_logger.log_excel_imported(run_id, result.get("stats", {}))
    if result.get("qc_issue_count"):
        audit_logger.log(
            operation_type="DATA_IMPORT",
            operation_name="导入数据异常",
            run_id=run_id,
            status="FAILED",
            error_message="导入数据存在异常",
            payload={
                "qc_issue_count": result.get("qc_issue_count"),
                "unmatched_items": result.get("unmatched_items"),
            },
        )
    return result


@router.post("/runs/{run_id}/import/excel/sheet")
def import_excel_sheet_api(
    run_id: int,
    sheet: str,
    file: UploadFile = File(...),
    db=Depends(get_db),
):
    run = db.get(RunBatch, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    model = RAW_SHEET_MAP.get(sheet)
    if model is None:
        raise HTTPException(status_code=404, detail="Unknown sheet")
    audit_logger = get_audit_logger(db)
    db.execute(delete(model).where(model.run_id == run_id))
    db.commit()
    tmp_path = None
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=TMP_DIR) as tmp:
        content = file.file.read()
        tmp.write(content)
        tmp_path = tmp.name
    try:
        result = import_excel(db, run_id, tmp_path, target_sheets={RAW_SHEET_LABELS.get(sheet, sheet)})
    except Exception as exc:
        audit_logger.log_error("DATA_IMPORT", "导入单表数据", str(exc), run_id=run_id)
        raise
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    result["cleared_existing"] = True
    result["sheet"] = sheet
    audit_logger.log(
        operation_type="DATA_IMPORT",
        operation_name="导入单表数据",
        run_id=run_id,
        details=f"导入 {RAW_SHEET_LABELS.get(sheet, sheet)}",
        payload=result,
    )
    if result.get("qc_issue_count"):
        audit_logger.log(
            operation_type="DATA_IMPORT",
            operation_name="导入数据异常",
            run_id=run_id,
            status="FAILED",
            error_message="导入数据存在异常",
            payload={
                "sheet": sheet,
                "qc_issue_count": result.get("qc_issue_count"),
                "unmatched_items": result.get("unmatched_items"),
            },
        )
    return result


@router.post("/runs/{run_id}/import/night-shift/schedules")
def import_night_shift_from_schedules(
    run_id: int,
    doctor_file: UploadFile = File(...),
    nurse_file: UploadFile = File(...),
    clear_existing: bool = Form(True),
    db=Depends(get_db),
):
    run = db.get(RunBatch, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    audit_logger = get_audit_logger(db)

    temp_paths: list[str] = []

    def _save_upload(upload: UploadFile) -> str:
        suffix = os.path.splitext(upload.filename or "")[1] or ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(upload.file.read())
            path = tmp.name
        temp_paths.append(path)
        return path

    try:
        doctor_path = _save_upload(doctor_file)
        nurse_path = _save_upload(nurse_file)
        doctor_counts = parse_night_shift_counts(doctor_path, role="doctor")
        nurse_counts = parse_night_shift_counts(nurse_path, role="nurse")
    except Exception as exc:
        audit_logger.log_error(
            "DATA_IMPORT",
            "按排班表计算夜班",
            str(exc),
            run_id=run_id,
        )
        raise HTTPException(status_code=400, detail=f"排班表解析失败: {exc}") from exc
    finally:
        for path in temp_paths:
            try:
                os.unlink(path)
            except Exception:
                pass

    if clear_existing:
        db.execute(delete(RawNightShift).where(RawNightShift.run_id == run_id))

    merged_counts: dict[str, Decimal] = {}
    for source in (doctor_counts, nurse_counts):
        for name, count in source.items():
            merged_counts[name] = merged_counts.get(name, Decimal("0")) + count

    for idx, (name, count) in enumerate(sorted(merged_counts.items()), start=2):
        db.add(
            RawNightShift(
                run_id=run_id,
                name=name,
                night_count=count,
                row_no=idx,
                sheet_name="排班表自动统计",
            )
        )
    db.commit()

    doctor_total = sum(doctor_counts.values(), Decimal("0"))
    nurse_total = sum(nurse_counts.values(), Decimal("0"))

    result = {
        "status": "ok",
        "cleared_existing": clear_existing,
        "doctor_rows": len(doctor_counts),
        "doctor_night_total": float(doctor_total),
        "nurse_rows": len(nurse_counts),
        "nurse_night_total": float(nurse_total),
        "imported_rows": len(merged_counts),
    }
    audit_logger.log(
        operation_type="DATA_IMPORT",
        operation_name="按排班表计算夜班",
        run_id=run_id,
        details=f"导入夜班统计，共 {len(merged_counts)} 人",
        payload=result,
    )
    return result


@router.post("/runs/{run_id}/validate")
def validate_run(run_id: int, save_to_qc: bool = True, db=Depends(get_db)):
    """验证 run 数据的完整性和合理性"""
    run = db.get(RunBatch, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    
    validation_result = validate_run_data(db, run_id)
    
    if save_to_qc:
        save_validation_results_to_qc(db, run_id, validation_result)
    
    return validation_result.to_dict()


@router.post("/runs/{run_id}/calculate")
def calculate_api(run_id: int, db=Depends(get_db)):
    run = db.get(RunBatch, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.status == "LOCKED":
        raise HTTPException(status_code=400, detail="Cannot calculate locked run")
    audit_logger = get_audit_logger(db)
    try:
        result = calculate_run(db, run_id)
        run.status = "CALCULATED"
        db.commit()
        audit_logger.log_calculation(run_id, result)
        return result
    except Exception as exc:
        audit_logger.log_error("CALCULATION", "执行绩效计算", str(exc), run_id=run_id)
        raise


@router.get("/runs/{run_id}/summary", response_model=SummaryResponse)
def get_summary(run_id: int, db=Depends(get_db)):
    rows = db.execute(
        select(FactPaySummary).where(FactPaySummary.run_id == run_id)
    ).scalars()
    return {
        "rows": [
            {
                "name": r.name,
                "role": r.role,
                "direct_total": float(r.direct_total),
                "pool_nursing": float(r.pool_nursing),
                "pool_doctor": float(r.pool_doctor),
                "surplus": float(r.surplus),
                "grand_total": float(r.grand_total),
            }
            for r in rows
        ]
    }


@router.get("/runs/{run_id}/detail")
def get_detail(run_id: int, name: Optional[str] = None, db=Depends(get_db)):
    stmt = select(FactPayDetail).where(FactPayDetail.run_id == run_id)
    if name:
        stmt = stmt.where(FactPayDetail.name == name)
    rows = db.execute(stmt).scalars().all()
    return [
        {
            "name": r.name,
            "item_code": r.item_code,
            "item_name": r.item_name,
            "amount": float(r.amount),
            "pay_type": r.pay_type,
            "pool_code": r.pool_code,
            "source_item_code": r.source_item_code,
            "note": r.calc_note,
        }
        for r in rows
    ]


@router.get("/runs/{run_id}/raw/{sheet}")
def get_raw_sheet(run_id: int, sheet: str, db=Depends(get_db)):
    model = RAW_SHEET_MAP.get(sheet)
    if model is None:
        raise HTTPException(status_code=404, detail="Unknown sheet")
    rows = db.execute(select(model).where(model.run_id == run_id)).scalars().all()
    def _to_dict(row):
        return {k: v for k, v in row.__dict__.items() if not k.startswith("_")}

    return [_to_dict(row) for row in rows]


@router.post("/runs/{run_id}/raw/{sheet}")
def save_raw_sheet(run_id: int, sheet: str, payload: dict, db=Depends(get_db)):
    model = RAW_SHEET_MAP.get(sheet)
    if model is None:
        raise HTTPException(status_code=404, detail="Unknown sheet")
    rows = payload.get("rows")
    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail="Invalid rows")
    audit_logger = get_audit_logger(db)
    db.execute(delete(model).where(model.run_id == run_id))
    sheet_name = RAW_SHEET_LABELS.get(sheet, sheet)
    for idx, row in enumerate(rows, start=2):
        if sheet == "hospital_perf_item":
            item_name = str(row.get("item_name") or "").strip()
            if not item_name:
                continue
            amount = to_decimal(row.get("amount"))
            if amount is None:
                continue
            item_name_norm = normalize_item_name(item_name)
            period_tag = extract_period_tag(item_name)
            db.add(
                RawHospitalPerfItem(
                    run_id=run_id,
                    item_name=item_name,
                    item_name_norm=item_name_norm,
                    period_tag=period_tag,
                    amount=amount,
                    row_no=idx,
                    sheet_name=sheet_name,
                )
            )
        elif sheet == "roster":
            name = str(row.get("name") or "").strip()
            if not name:
                continue
            role = str(row.get("role") or "")
            perf_score = to_decimal(row.get("perf_score"))
            eligible = row.get("eligible_for_surplus_weight")
            eligible_val = None if eligible is None else bool(eligible)
            db.add(
                RawRoster(
                    run_id=run_id,
                    name=name,
                    role=role,
                    perf_score=perf_score,
                    eligible_for_surplus_weight=eligible_val,
                    row_no=idx,
                    sheet_name=sheet_name,
                )
            )
        elif sheet == "night_shift":
            name = str(row.get("name") or "").strip()
            if not name:
                continue
            db.add(
                RawNightShift(
                    run_id=run_id,
                    name=name,
                    night_count=to_decimal(row.get("night_count")),
                    row_no=idx,
                    sheet_name=sheet_name,
                )
            )
        elif sheet == "reading_fee":
            name = str(row.get("name") or "").strip()
            if not name:
                continue
            db.add(
                RawReadingFee(
                    run_id=run_id,
                    category=str(row.get("category") or ""),
                    name=name,
                    amount=to_decimal(row.get("amount")),
                    row_no=idx,
                    sheet_name=sheet_name,
                )
            )
        elif sheet == "doctor_workload":
            name = str(row.get("name") or "").strip()
            if not name:
                continue
            db.add(
                RawDoctorWorkload(
                    run_id=run_id,
                    name=name,
                    workload=to_decimal(row.get("workload")),
                    bed_days=to_decimal(row.get("bed_days")),
                    admission_cert_count=to_decimal(row.get("admission_cert_count")),
                    row_no=idx,
                    sheet_name=sheet_name,
                )
            )
        elif sheet == "nurse_workload":
            name = str(row.get("name") or "").strip()
            if not name:
                continue
            db.add(
                RawNurseWorkload(
                    run_id=run_id,
                    name=name,
                    score=to_decimal(row.get("score")),
                    blood_draw_count=to_decimal(row.get("blood_draw_count")),
                    row_no=idx,
                    sheet_name=sheet_name,
                )
            )
        elif sheet == "manual_doctor_workload_pay":
            name = str(row.get("name") or "").strip()
            if not name:
                continue
            db.add(
                RawManualDoctorWorkloadPay(
                    run_id=run_id,
                    name=name,
                    amount=to_decimal(row.get("amount")),
                    row_no=idx,
                    sheet_name=sheet_name,
                )
            )
    db.commit()
    audit_logger.log(
        operation_type="DATA_EDIT",
        operation_name="编辑导入数据",
        run_id=run_id,
        details=f"编辑 {sheet_name}，共 {len(rows)} 条",
        payload={"sheet": sheet, "count": len(rows)},
    )
    return {"status": "ok", "count": len(rows)}


@router.get("/runs/{run_id}/reconcile")
def get_reconcile(run_id: int, db=Depends(get_db)):
    rows = db.execute(
        select(ReconcileItem).where(ReconcileItem.run_id == run_id)
    ).scalars().all()

    # 映射 item_code -> 原始项目名（来自项目映射表）
    mapping_rows = (
        db.execute(
            select(DictItemMapping).where(DictItemMapping.is_active == True)
        )
        .scalars()
        .all()
    )
    item_name_map = {m.item_code: m.raw_item_name for m in mapping_rows}
    manual_name_map = {
        "MANUAL_OTHER_PAY": "其他（手工录入）",
    }

    response = []
    for r in rows:
        item_name = item_name_map.get(r.item_code, manual_name_map.get(r.item_code, r.item_code))
        note = r.note
        if note and note.startswith("RAW_NAME:"):
            item_name = note.replace("RAW_NAME:", "", 1)
            note = ""
        response.append(
            {
                "item_code": r.item_code,
                # 前端对账界面希望看到“原始项目名”，因此这里返回映射表中的 raw_item_name。
                # 若找不到映射，则退回显示内部编码。
                "item_name": item_name,
                "source_amount": float(r.source_amount),
                "allocated_amount": float(r.allocated_amount),
                "delta": float(r.delta),
                "note": note,
            }
        )
    return response


@router.get("/runs/{run_id}/qc")
def get_qc(run_id: int, db=Depends(get_db)):
    rows = db.execute(select(QcIssue).where(QcIssue.run_id == run_id)).scalars().all()
    return [
        {
            "issue_type": r.issue_type,
            "message": r.message,
            "severity": r.severity,
            "payload": r.payload,
        }
        for r in rows
    ]


@router.get("/runs/{run_id}/export/excel")
def export_excel(run_id: int, db=Depends(get_db)):
    import openpyxl

    audit_logger = get_audit_logger(db)
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "汇总"
    ws_detail = wb.create_sheet("明细")
    ws_reconcile = wb.create_sheet("对账")
    ws_qc = wb.create_sheet("异常")

    # Summary
    ws_summary.append(["姓名", "岗位", "DirectPay合计", "护理池分配", "医师池分配", "最终应发合计"])
    for r in db.execute(select(FactPaySummary).where(FactPaySummary.run_id == run_id)).scalars():
        ws_summary.append(
            [
                r.name,
                r.role,
                float(r.direct_total),
                float(r.pool_nursing),
                float(r.pool_doctor),
                float(r.grand_total),
            ]
        )

    # Detail
    ws_detail.append(["姓名", "项目编码", "项目名", "金额", "类型", "池子", "来源编码"])
    for r in db.execute(select(FactPayDetail).where(FactPayDetail.run_id == run_id)).scalars():
        ws_detail.append(
            [
                r.name,
                r.item_code,
                r.item_name,
                float(r.amount),
                r.pay_type,
                r.pool_code,
                r.source_item_code,
            ]
        )

    # Reconcile
    ws_reconcile.append(["项目编码", "来源金额", "已分配", "差额", "备注"])
    for r in db.execute(select(ReconcileItem).where(ReconcileItem.run_id == run_id)).scalars():
        ws_reconcile.append(
            [
                r.item_code,
                float(r.source_amount),
                float(r.allocated_amount),
                float(r.delta),
                r.note,
            ]
        )

    # QC
    ws_qc.append(["类型", "消息", "严重级别", "Payload"])
    for r in db.execute(select(QcIssue).where(QcIssue.run_id == run_id)).scalars():
        ws_qc.append([r.issue_type, r.message, r.severity, str(r.payload)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    audit_logger.log_export(run_id, "excel")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=run_{run_id}_export.xlsx"},
    )


@router.get("/runs/{run_id}/export/pdf")
def export_pdf(
    run_id: int,
    paper: str = "A4",
    orientation: str = "landscape",
    # PDF 默认仅输出“汇总 + 签字页”（不包含对账/异常）。
    sections: str = "summary,sign",
    db=Depends(get_db),
):
    from app.reporting.pdf_renderer import PdfRenderParams, PdfRendererUnavailable, render_pdf_from_html
    from app.reporting.run_report import RunReportOptions, build_run_report_context, render_run_report_html

    try:
        options = RunReportOptions(
            paper=paper,
            orientation=orientation.lower(),
            sections=tuple([s.strip().lower() for s in sections.split(",") if s.strip()]),
        )
        context = build_run_report_context(
            db=db,
            run_id=run_id,
            hospital_name=PDF_HOSPITAL_NAME,
            sections=options.sections,
        )
    except KeyError:
        raise HTTPException(status_code=404, detail="Run not found")

    html = render_run_report_html(context=context, options=options)
    try:
        pdf_bytes = render_pdf_from_html(
            html=html,
            params=PdfRenderParams(paper=options.paper, landscape=(options.orientation == "landscape")),
        )
    except PdfRendererUnavailable as e:
        raise HTTPException(status_code=503, detail=str(e))

    get_audit_logger(db).log_export(run_id, "pdf")
    buf = io.BytesIO(pdf_bytes)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=run_{run_id}_export.pdf"},
    )


@router.get("/runs/{run_id}/export/html")
def export_html(
    run_id: int,
    paper: str = "A4",
    orientation: str = "landscape",
    # HTML 默认保留完整信息，便于预览与调样式（需要的话可通过 sections 裁剪）。
    sections: str = "summary,reconcile,qc,sign",
    db=Depends(get_db),
):
    from fastapi.responses import HTMLResponse

    from app.reporting.run_report import RunReportOptions, build_run_report_context, render_run_report_html

    try:
        options = RunReportOptions(
            paper=paper,
            orientation=orientation.lower(),
            sections=tuple([s.strip().lower() for s in sections.split(",") if s.strip()]),
        )
        context = build_run_report_context(
            db=db,
            run_id=run_id,
            hospital_name=PDF_HOSPITAL_NAME,
            sections=options.sections,
        )
    except KeyError:
        raise HTTPException(status_code=404, detail="Run not found")

    html = render_run_report_html(context=context, options=options)
    get_audit_logger(db).log_export(run_id, "html")
    return HTMLResponse(content=html)


@router.get("/runs/compare")
def compare_runs(run_id_1: int, run_id_2: int, db=Depends(get_db)):
    """对比两个批次的计算结果"""
    run1 = db.get(RunBatch, run_id_1)
    run2 = db.get(RunBatch, run_id_2)
    
    if not run1 or not run2:
        raise HTTPException(status_code=404, detail="One or both runs not found")
    
    # 获取两个批次的汇总数据
    summary1 = {
        r.name: r for r in db.execute(
            select(FactPaySummary).where(FactPaySummary.run_id == run_id_1)
        ).scalars()
    }
    summary2 = {
        r.name: r for r in db.execute(
            select(FactPaySummary).where(FactPaySummary.run_id == run_id_2)
        ).scalars()
    }
    
    # 对比结果
    all_names = set(summary1.keys()) | set(summary2.keys())
    comparison = []
    
    for name in sorted(all_names):
        s1 = summary1.get(name)
        s2 = summary2.get(name)
        
        if s1 and s2:
            comparison.append({
                "name": name,
                "role": s1.role,
                "run1_total": float(s1.grand_total),
                "run2_total": float(s2.grand_total),
                "diff": float(s2.grand_total - s1.grand_total),
                "diff_percent": float((s2.grand_total - s1.grand_total) / s1.grand_total * 100) if s1.grand_total != 0 else 0,
                "status": "both"
            })
        elif s1:
            comparison.append({
                "name": name,
                "role": s1.role,
                "run1_total": float(s1.grand_total),
                "run2_total": 0,
                "diff": float(-s1.grand_total),
                "diff_percent": -100,
                "status": "only_in_run1"
            })
        else:
            comparison.append({
                "name": name,
                "role": s2.role,
                "run1_total": 0,
                "run2_total": float(s2.grand_total),
                "diff": float(s2.grand_total),
                "diff_percent": 100,
                "status": "only_in_run2"
            })
    
    return {
        "run1": {"id": run_id_1, "month": run1.month, "dept_name": run1.dept_name},
        "run2": {"id": run_id_2, "month": run2.month, "dept_name": run2.dept_name},
        "comparison": comparison,
        "summary": {
            "total_people": len(all_names),
            "common_people": len(set(summary1.keys()) & set(summary2.keys())),
            "only_in_run1": len(set(summary1.keys()) - set(summary2.keys())),
            "only_in_run2": len(set(summary2.keys()) - set(summary1.keys())),
        }
    }


@router.post("/runs/{run_id}/copy")
def copy_run(run_id: int, new_month: str, db=Depends(get_db)):
    """复制一个批次（仅复制配置，不复制数据）"""
    source_run = db.get(RunBatch, run_id)
    if not source_run:
        raise HTTPException(status_code=404, detail="Source run not found")
    
    new_run = RunBatch(
        month=new_month,
        dept_name=source_run.dept_name,
        rule_version=source_run.rule_version,
        status="DRAFT"
    )
    db.add(new_run)
    db.commit()
    db.refresh(new_run)
    
    return {
        "run_id": new_run.id,
        "month": new_run.month,
        "dept_name": new_run.dept_name,
        "rule_version": new_run.rule_version,
        "copied_from": run_id
    }


@router.delete("/runs/{run_id}")
def delete_run(run_id: int, db=Depends(get_db)):
    """删除批次及其所有相关数据"""
    run = db.get(RunBatch, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    
    # 检查是否已锁定
    if run.status == "LOCKED":
        raise HTTPException(status_code=400, detail="Cannot delete locked run")
    
    # 删除所有相关数据
    _delete_run_rows(db, run_id, RUN_SCOPED_MODELS)

    # 删除批次本身
    db.delete(run)
    db.commit()
    get_audit_logger(db).log(
        operation_type="RUN_MANAGEMENT",
        operation_name="删除核算批次",
        run_id=run_id,
        details=f"删除核算批次 {run_id}",
    )
    
    return {"status": "deleted", "run_id": run_id}
