from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

XLS_SIGNATURE = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

DOCTOR_NIGHT_EXACT = {
    "夜",
    "24小时",
}
DOCTOR_NIGHT_CONTAINS = {
    "夜",
}

NURSE_NIGHT_EXACT = {
    "夜",
    "上夜",
    "上大",
    "大",
    "小",
    "通夜",
    "内二科通夜",
}
NURSE_NIGHT_CONTAINS = {
    "夜",
    "通夜",
}


def parse_night_shift_counts(file_path: str, role: str) -> dict[str, Decimal]:
    role = role.strip().lower()
    if role not in {"doctor", "nurse"}:
        raise ValueError(f"Unsupported role: {role}")

    rows = _load_schedule_rows(file_path)
    header_row_idx, name_col_idx = _find_name_header(rows)
    if header_row_idx is None or name_col_idx is None:
        raise ValueError("排班表中未找到“姓名”表头")

    counts = defaultdict(lambda: Decimal("0"))
    for row in rows[header_row_idx + 1 :]:
        name = _clean_text(_get_cell(row, name_col_idx))
        if not name:
            continue
        night_count = Decimal("0")
        for raw_shift in row[name_col_idx + 1 :]:
            if _is_night_shift(raw_shift, role):
                night_count += Decimal("1")
        if night_count > 0:
            counts[name] += night_count

    return dict(counts)


def _load_schedule_rows(file_path: str) -> list[list[Any]]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".xls" or _is_xls_file(path):
        return _load_schedule_rows_xls(str(path))
    return _load_schedule_rows_xlsx(str(path))


def _is_xls_file(path: Path) -> bool:
    try:
        with path.open("rb") as fp:
            return fp.read(8) == XLS_SIGNATURE
    except Exception:
        return False


def _load_schedule_rows_xlsx(file_path: str) -> list[list[Any]]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _load_schedule_rows_xls(file_path: str) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("读取 .xls 需要安装 xlrd") from exc

    wb = xlrd.open_workbook(file_path)
    sheet = wb.sheet_by_index(0)
    return [sheet.row_values(i) for i in range(sheet.nrows)]


def _find_name_header(rows: list[list[Any]]) -> tuple[int | None, int | None]:
    max_scan_rows = min(30, len(rows))
    for row_idx in range(max_scan_rows):
        row = rows[row_idx]
        for col_idx, cell in enumerate(row):
            if _clean_text(cell) == "姓名":
                return row_idx, col_idx
    return None, None


def _get_cell(row: list[Any], idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).replace("\r", "").replace("\n", "").strip()


def _is_night_shift(raw_shift: Any, role: str) -> bool:
    shift = _clean_text(raw_shift).replace(" ", "")
    if not shift:
        return False

    if role == "doctor":
        exact = DOCTOR_NIGHT_EXACT
        contains = DOCTOR_NIGHT_CONTAINS
    else:
        exact = NURSE_NIGHT_EXACT
        contains = NURSE_NIGHT_CONTAINS

    if shift in exact:
        return True
    return any(token in shift for token in contains)
