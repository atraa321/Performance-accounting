from __future__ import annotations

from decimal import Decimal
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import select

from app.calc.utils import normalize_item_name, extract_period_tag, to_decimal
from app.models.models import (
    RawHospitalPerfItem,
    RawRoster,
    RawNightShift,
    RawReadingFee,
    RawDoctorWorkload,
    RawNurseWorkload,
    RawManualDoctorWorkloadPay,
    RawManualPoolAdjust,
    DictItemMapping,
    QcIssue,
)


def _header_map(headers: list[str]) -> dict[str, int]:
    mapping = {}
    for idx, col in enumerate(headers):
        if col is None:
            continue
        key = str(col).strip()
        mapping[key] = idx
    return mapping


def _find_col(mapping: dict[str, int], *candidates: str) -> Optional[int]:
    for name in candidates:
        if name in mapping:
            return mapping[name]
    return None


def import_excel(session, run_id: int, file_path: str, target_sheets: Optional[set[str]] = None) -> dict[str, Any]:
    wb = load_workbook(file_path, data_only=True)

    stats = {}
    qc_issues = []

    for sheet_name in wb.sheetnames:
        if target_sheets and sheet_name not in target_sheets:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_row = rows[0]
        header_map = _header_map(list(header_row))

        if sheet_name in ("院发绩效表",):
            col_name = _find_col(header_map, "项目名", "科室", "项目")
            col_amount = _find_col(header_map, "金额")
            if col_name is None or col_amount is None:
                continue
            count = 0
            # 获取数据行（排除表头）
            data_rows = rows[1:]
            for row_no, row in enumerate(data_rows, start=2):
                item_name = row[col_name]
                amount = row[col_amount]
                
                # 跳过空行
                if item_name is None:
                    continue
                
                # 跳过最后一行（通常是合计行）
                if row_no == len(data_rows) + 1:
                    continue
                
                # 检查是否是合计行（通过项目名判断）
                item_name_str = str(item_name).strip()
                if item_name_str in ("合计", "总计", "小计", "汇总"):
                    continue
                
                # 转换金额
                amount_decimal = to_decimal(amount)
                
                # 排除金额为0的项目
                if amount_decimal is None or amount_decimal == 0:
                    continue
                
                item_name_norm = normalize_item_name(item_name_str)
                period_tag = extract_period_tag(item_name_str)
                session.add(
                    RawHospitalPerfItem(
                        run_id=run_id,
                        item_name=item_name_str,
                        item_name_norm=item_name_norm,
                        period_tag=period_tag,
                        amount=amount_decimal,
                        row_no=row_no,
                        sheet_name=sheet_name,
                    )
                )
                count += 1
            stats[sheet_name] = count

        elif sheet_name in ("绩效发放名单",):
            col_name = _find_col(header_map, "姓名")
            col_role = _find_col(header_map, "岗位")
            col_score = _find_col(header_map, "绩效分数", "绩效")
            col_eligible = _find_col(header_map, "是否参与盈余", "盈余参与")
            if col_name is None or col_role is None or col_score is None:
                continue
            count = 0
            for row_no, row in enumerate(rows[1:], start=2):
                name = row[col_name]
                if name is None:
                    continue
                role = row[col_role]
                score = row[col_score]
                eligible = None
                if col_eligible is not None:
                    eligible = bool(row[col_eligible]) if row[col_eligible] is not None else None
                session.add(
                    RawRoster(
                        run_id=run_id,
                        name=str(name),
                        role=str(role or ""),
                        perf_score=to_decimal(score),
                        eligible_for_surplus_weight=eligible,
                        row_no=row_no,
                        sheet_name=sheet_name,
                    )
                )
                count += 1
            stats[sheet_name] = count

        elif sheet_name in ("夜班统计",):
            def _add_night_row(row_no, name_cell, count_cell):
                if name_cell is None:
                    return 0
                session.add(
                    RawNightShift(
                        run_id=run_id,
                        name=str(name_cell),
                        night_count=to_decimal(count_cell),
                        row_no=row_no,
                        sheet_name=sheet_name,
                    )
                )
                return 1

            # 仅支持新模板：姓名 + 夜班数 + 岗位
            col_name = _find_col(header_map, "姓名")
            col_count = _find_col(header_map, "夜班数", "夜班")
            col_role = _find_col(header_map, "岗位")
            if col_name is None or col_count is None or col_role is None:
                continue

            count = 0
            for row_no, row in enumerate(rows[1:], start=2):
                count += _add_night_row(row_no, row[col_name], row[col_count])
            stats[sheet_name] = count

        elif sheet_name in ("判读费",):
            col_category = _find_col(header_map, "类别", "判读类别")
            col_name = _find_col(header_map, "姓名")
            col_amount = _find_col(header_map, "金额", "求和项:判读金额")
            if col_category is None or col_name is None or col_amount is None:
                continue
            count = 0
            last_category = None
            for row_no, row in enumerate(rows[1:], start=2):
                category = row[col_category]
                amount = row[col_amount]
                if category is not None and str(category).strip():
                    last_category = category
                elif last_category is not None:
                    category = last_category
                name = row[col_name]
                if name is None:
                    continue
                session.add(
                    RawReadingFee(
                        run_id=run_id,
                        category=str(category or ""),
                        name=str(name),
                        amount=to_decimal(amount),
                        row_no=row_no,
                        sheet_name=sheet_name,
                    )
                )
                count += 1
            stats[sheet_name] = count

        elif sheet_name in ("医师工作量",):
            col_name = _find_col(header_map, "姓名", "医师")
            col_workload = _find_col(header_map, "工作量")
            col_bed = _find_col(header_map, "床日数")
            col_admission = _find_col(header_map, "住院证", "住院证数")
            if col_name is None or col_workload is None or col_bed is None or col_admission is None:
                continue
            count = 0
            for row_no, row in enumerate(rows[1:], start=2):
                name = row[col_name]
                if name is None:
                    continue
                session.add(
                    RawDoctorWorkload(
                        run_id=run_id,
                        name=str(name),
                        workload=to_decimal(row[col_workload]),
                        bed_days=to_decimal(row[col_bed]),
                        admission_cert_count=to_decimal(row[col_admission]),
                        row_no=row_no,
                        sheet_name=sheet_name,
                    )
                )
                count += 1
            stats[sheet_name] = count

        elif sheet_name in ("护士工作量",):
            col_name = _find_col(header_map, "姓名")
            col_score = _find_col(header_map, "分数")
            col_blood = _find_col(header_map, "抽血量")
            if col_name is None or col_score is None or col_blood is None:
                continue
            count = 0
            for row_no, row in enumerate(rows[1:], start=2):
                name = row[col_name]
                if name is None:
                    continue
                session.add(
                    RawNurseWorkload(
                        run_id=run_id,
                        name=str(name),
                        score=to_decimal(row[col_score]),
                        blood_draw_count=to_decimal(row[col_blood]),
                        row_no=row_no,
                        sheet_name=sheet_name,
                    )
                )
                count += 1
            stats[sheet_name] = count

        elif sheet_name in ("医师手动工作量分配",):
            col_name = _find_col(header_map, "姓名")
            col_amount = _find_col(header_map, "金额")
            if col_name is None or col_amount is None:
                continue
            count = 0
            for row_no, row in enumerate(rows[1:], start=2):
                name = row[col_name]
                if name is None:
                    continue
                session.add(
                    RawManualDoctorWorkloadPay(
                        run_id=run_id,
                        name=str(name),
                        amount=to_decimal(row[col_amount]),
                        row_no=row_no,
                        sheet_name=sheet_name,
                    )
                )
                count += 1
            stats[sheet_name] = count

        elif sheet_name in ("手动池调整", "手动池补录"):
            col_pool = _find_col(header_map, "池子", "池", "pool_code")
            col_amount = _find_col(header_map, "金额")
            if col_pool is None or col_amount is None:
                continue
            count = 0
            for row_no, row in enumerate(rows[1:], start=2):
                pool_code = row[col_pool]
                if pool_code is None:
                    continue
                session.add(
                    RawManualPoolAdjust(
                        run_id=run_id,
                        pool_code=str(pool_code),
                        amount=to_decimal(row[col_amount]),
                        note=None,
                        row_no=row_no,
                        sheet_name=sheet_name,
                    )
                )
                count += 1
            stats[sheet_name] = count

    session.commit()

    # Detect unmapped items
    mapping_rows = session.execute(select(DictItemMapping).where(DictItemMapping.is_active == True)).scalars()
    mapping = [m.raw_item_name for m in mapping_rows]
    unmatched = []
    for item in session.execute(
        select(RawHospitalPerfItem).where(RawHospitalPerfItem.run_id == run_id)
    ).scalars():
        norm = item.item_name_norm
        if not any(key in norm for key in mapping):
            unmatched.append(item.item_name)
            qc_issues.append(
                QcIssue(
                    run_id=run_id,
                    issue_type="ITEM_MAPPING_MISSING",
                    message=f"Missing mapping for item: {item.item_name}",
                    severity="WARN",
                    payload={"item_name": item.item_name},
                    sheet_name=item.sheet_name,
                    row_no=item.row_no,
                )
            )

    if qc_issues:
        session.add_all(qc_issues)
        session.commit()

    return {
        "stats": stats,
        "unmatched_items": sorted(set(unmatched)),
        "qc_issue_count": len(qc_issues),
    }
